import ast
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import time
from datetime import datetime
from copy import copy
import os 

nfOrderpath = "C:\\Users\\amol37007\\OneDrive\\JsonTxt\\Current\\NiftyOrders.txt"
grOrderpath = "C:\\Users\\amol37007\\OneDrive\\JsonTxt\\Current\\GROrders.txt"

KHexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\KHlog.xlsx"
BYexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\BYlog.xlsx"
# VHexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\VHlog.xlsx"
HVexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\HVlog.xlsx"
# AKexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\AKlog.xlsx"


############### Excel loading and writing #############

def log_to_excel(excelpath,logList):	

	wb = load_workbook(excelpath)
	ws = wb.worksheets[0]

	lastrow = ws.max_row
	print("lastrow:")
	print(lastrow)

	sl =  ws.max_row -2

	ws.insert_rows( ws.max_row -1)

	newRowLocation = ws.max_row -2
	print("newRowLocation:")
	print(newRowLocation)

	cell_obj = ws.cell(row=sl, column=1)
	print(type(cell_obj))
	lastTradeNo = cell_obj.value
	print("Tr.No : " +str(lastTradeNo))

	#logList2 = [tradeType,trdate,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts,qty,pnl,strategy]

	# #write to the cell you want, specifying row and column, and value :-)
	ws.cell(column=1,row=newRowLocation, value=(lastTradeNo+1))     #Tr.no	

	tradetype = ws.cell(column=3,row=newRowLocation)			 #tradeType
	tradetype.value = logList[0]
	if tradetype.value == 'Short':
		tradetype.fill = PatternFill('solid',fgColor='F4B084')
		tradetype.font = Font(italic = True) 
	else:
		tradetype.fill = PatternFill('solid',fgColor='A9D08E')
		tradetype.font = Font(italic = True)

	trDate = ws.cell(column=4,row=newRowLocation)       #date
	trDate.value = logList[1]

	entryTime = ws.cell(column=5,row=newRowLocation)			 #entryTime
	entryTime.fill =  PatternFill('solid',fgColor='D6DCE4')
	entryTime.value = logList[2]

	exitTime = ws.cell(column=6,row=newRowLocation)				 #exitTime
	exitTime.fill =  PatternFill('solid',fgColor='D6DCE4') 
	exitTime.value = logList[3]
	
	ws.cell(column=7,row=newRowLocation, value=logList[4])     	 #entry price
	ws.cell(column=8,row=newRowLocation, value=logList[5])       #exit price
	ws.cell(column=9,row=newRowLocation, value=logList[6])       #hedge cost

	tradePoints = ws.cell(column=10,row=newRowLocation)			 #TradePts
	tradePoints.value = logList[7]
	if tradePoints.value <=0 :
		tradePoints.fill = PatternFill('solid',fgColor='FFC7CE')
		tradePoints.font = Font(size = 14,bold = True,color = '9C0006')
	else:
		tradePoints.fill =PatternFill('solid',fgColor='C6EFCE')
		tradePoints.font = Font(size = 14,bold = True,color = '006100')

	qty = ws.cell(column=11,row=newRowLocation)					 #qty
	qty.fill = PatternFill('solid',fgColor='638EC6')
	qty.value = logList[8]

	PnL = ws.cell(column=12,row=newRowLocation)        			 #PnL
	PnL.value = logList[9]
	if PnL.value <= 0 :
		PnL.fill = PatternFill('solid',fgColor='FFC7CE')
		PnL.font = Font(size = 14,bold = True,color = '9C0006')
	else:
		PnL.fill =PatternFill('solid',fgColor='C6EFCE')
		PnL.font = Font(size = 14,bold = True,color = '006100')
	
	strategy =  ws.cell(column=2,row=newRowLocation)			 #strategy
	if logList[10] == "Golden Ratio" :
		strategy.value = "Golden Ratio"
		strategy.fill =  PatternFill('solid',fgColor='FFD966')
		strategy.font = Font(color = '9C0006')
	else:
		strategy.value = "Nifty Straddle"
		strategy.fill = PatternFill('solid',fgColor='B4C6E7') 

	totalPnL =  ws.cell(column=12,row=ws.max_row)				#totalPnL
	totalPnL.value ='=sum(L3:L{})'.format(ws.max_row-2)
	
	thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	for row in ws[newRowLocation:newRowLocation]:  # skip the header
		row.alignment = Alignment(horizontal='center')
		row.border = thin_border

	wb.save(filename=excelpath)
	wb.close()

def shortTradeEntry(shrtOrdrData,excelFile):
	tradeType = 'Short'	
	strategy = 'Nifty Straddle'
	
	dt_object = shrtOrdrData[1]["timestamp"]
	entryTimeStamp = datetime.strptime(dt_object, '%d/%m/%Y %H:%M:%S')
	
	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")
	year = entryTimeStamp.strftime("%Y")

	trdate = (day+"-"+month+"-"+year)

	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = shrtOrdrData[2]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)

	entry_prc = shrtOrdrData[1]["exe_prc"]
	exit_prc = shrtOrdrData[2]["exe_prc"]
	hedge_cost = shrtOrdrData[0]["exe_prc"] - shrtOrdrData[3]["exe_prc"]
	trade_pts = entry_prc - exit_prc - hedge_cost
	qty = shrtOrdrData[1]["qty"]
	pnl = trade_pts * qty

	logList = [tradeType,trdate,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts,qty,pnl,strategy]

	log_to_excel(excelFile,logList)

def longTradeEntry(longOrdrData,excelFile):
	tradeType = 'Long'
	strategy = 'Nifty Straddle'	
	
	dt_object = longOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, "%d/%m/%Y %H:%M:%S")

	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")
	year = entryTimeStamp.strftime("%Y")

	trdate = (day+"-"+month+"-"+year)

	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = longOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)

	entry_prc = longOrdrData[0]["exe_prc"]
	exit_prc = longOrdrData[1]["exe_prc"]
	hedge_cost = 0.0
	trade_pts = exit_prc - entry_prc
	qty = longOrdrData[1]["qty"]
	pnl = trade_pts * qty

	logList = [tradeType,trdate,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts,qty,pnl,strategy]

	log_to_excel(excelFile,logList)	
	
def grTradeEntry(shrtOrdrData,excelFile,tradeType):
	strategy = 'Golden Ratio'
	
	dt_object = shrtOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, '%d/%m/%Y %H:%M:%S')
	
	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")
	year = entryTimeStamp.strftime("%Y")

	trdate = (day+"-"+month+"-"+year)

	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = shrtOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)

	entry_prc = shrtOrdrData[0]["prc"]
	exit_prc = shrtOrdrData[1]["prc"]
	hedge_cost = 0
	trade_pts = exit_prc - entry_prc 
	qty = shrtOrdrData[0]["qty"]
	pnl = trade_pts * qty

	logList = [tradeType,trdate,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts,qty,pnl,strategy]

	log_to_excel(excelFile,logList)


def settle(user_data,logbook):
	for order in range(len(user_data)-1):
		if user_data[order]["order_type"] == "HedgeEntry":
			shortTradeEntry(user_data[order:order+4],logbook)	

		if user_data[order]["order_type"] == "long":
			longTradeEntry(user_data[order:order+2],logbook)

		if user_data[order]["order_type"] == "golden_long":
			grTradeEntry(user_data[order:order+2],logbook,'Long')   # should give stratergy name 

		if user_data[order]["order_type"] == "golden_short":
			grTradeEntry(user_data[order:order+2],logbook,'Short')

#####################################################################

def straddleLog():
	with open(nfOrderpath, 'r') as f2:
		data = f2.read()

	data  = data.replace("\n","").split("]")
	d2 = []

	kh = []
	by = []
	vh = []
	hv = []
	ak = []

	for i in range(0,len(data)-1):
		data[i] = data[i].strip().replace("[","")
		data[i] = " ".join(data[i].split())
		data[i] = ast.literal_eval(data[i])
		d2.append(data[i])
		
	for d in d2:
		for x in d:
			if x['acc_name'] == "KH":
				kh.append(x)			
			elif x['acc_name'] == "BY":
				by.append(x)			
			elif x['acc_name'] == "VH":
				vh.append(x)			
			elif x['acc_name'] == "HV":
				hv.append(x)			
			elif x['acc_name'] == "AK":
				ak.append(x)

	settle(kh,KHexcelpath)
	settle(by,BYexcelpath)
	# settle(vh,VHexcelpath)
	settle(hv,HVexcelpath)
	# settle(ak,AKexcelpath)

def goldenLog():
	with open(grOrderpath, 'r') as f2:
		data = f2.read()

	data  = data.replace("\n","").split("]")
	d2 = []

	kh = []
	by = []
	vh = []
	hv = []
	ak = []

	for i in range(0,len(data)-1):
		data[i] = data[i].strip().replace("[","")
		data[i] = " ".join(data[i].split())
		data[i] = ast.literal_eval(data[i])
		d2.append(data[i])
	
	for d in d2:
		for x in d:
			if x['acc_name'] == "KH":
				kh.append(x)			
			elif x['acc_name'] == "BY":
				by.append(x)			
			elif x['acc_name'] == "VH":
				vh.append(x)			
			elif x['acc_name'] == "HV":
				hv.append(x)			
			elif x['acc_name'] == "AK":
				ak.append(x)

	settle(kh,KHexcelpath)
	settle(by,BYexcelpath)
	# settle(vh,VHexcelpath)
	settle(hv,HVexcelpath)
	# settle(ak,AKexcelpath)

def checkFile():
	if os.path.isfile(nfOrderpath):
		straddleLog()
	else:
		print ("No Trades today !")

	if os.path.isfile(grOrderpath):
		goldenLog()
	else:
		print ("No Golden Ratio Trades today !")

checkFile()

