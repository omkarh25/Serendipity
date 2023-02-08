import ast
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import time
from datetime import datetime
from copy import copy
import os 

tradelogpath = "C:\\Users\\amol37007\\OneDrive\\JsonTxt\\Current\\tradeLog.txt"
grlogpath = "C:\\Users\\amol37007\\OneDrive\\JsonTxt\\Current\\goldenLog.txt"

logexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\Master Trade log Signal.xlsx"

def log_to_excel(excelpath,logList):	

	wb = load_workbook(excelpath)
	ws = wb.worksheets[0]

	lastrow = ws.max_row
	print("lastrow:")
	print(lastrow)

	sl =  ws.max_row -4

	ws.insert_rows( ws.max_row -3)

	newRowLocation = ws.max_row -4
	print("newRowLocation:")
	print(newRowLocation)

	cell_obj = ws.cell(row=sl, column=1)
	print(type(cell_obj))
	lastTradeNo = cell_obj.value
	print("Tr.No : " +str(lastTradeNo))

# logList = [tradeType,strPrc,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts]

	# #write to the cell you want, specifying row and column, and value :-)
	ws.cell(column=1,row=newRowLocation, value=(lastTradeNo+1))     #Tr.no

	tradetype = ws.cell(column=2,row=newRowLocation)			 #tradeType
	tradetype.value = logList[0]
	if tradetype.value == 'Short':
		tradetype.fill = PatternFill('solid',fgColor='F4B084')
		tradetype.font = Font(italic = True) 
	else:
		tradetype.fill = PatternFill('solid',fgColor='A9D08E')
		tradetype.font = Font(italic = True)

	ws.cell(column=3,row=newRowLocation, value=logList[1])				#strPrc

	ws.cell(column=4,row=newRowLocation, value=logList[2])      	 #date
	ws.cell(column=5,row=newRowLocation, value=logList[3]) 			#day

	entryTime = ws.cell(column=6,row=newRowLocation)			 #entryTime
	entryTime.fill =  PatternFill('solid',fgColor='D6DCE4')
	entryTime.value = logList[4]

	exitTime = ws.cell(column=7,row=newRowLocation)				 #exitTime
	exitTime.fill =  PatternFill('solid',fgColor='D6DCE4') 
	exitTime.value = logList[5]
	
	ws.cell(column=8,row=newRowLocation, value=logList[6])     	 #entry price
	ws.cell(column=9,row=newRowLocation, value=logList[7])       #exit price
	ws.cell(column=10,row=newRowLocation, value=logList[8]) 		#hedge cost

	tradePoints = ws.cell(column=11,row=newRowLocation)			 #TradePts
	tradePoints.value = logList[9]
	if tradePoints.value <=0 :
		tradePoints.fill = PatternFill('solid',fgColor='FFC7CE')
		tradePoints.font = Font(size = 14,bold = True,color = '9C0006')
	else:
		tradePoints.fill =PatternFill('solid',fgColor='C6EFCE')
		tradePoints.font = Font(size = 14,bold = True,color = '006100')


	totalPnL =  ws.cell(column=11,row=ws.max_row)				#totalPnL
	totalPnL.value ='=sum(L3:L{})'.format(ws.max_row-4)
	
	thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	for row in ws[newRowLocation:newRowLocation]:  # skip the header
		row.alignment = Alignment(horizontal='center')
		row.border = thin_border

	wb.save(filename=excelpath)
	wb.close()

def gr_log_to_excel(excelpath,logList):	

	wb = load_workbook(excelpath)
	ws = wb.worksheets[1]

	lastrow = ws.max_row
	print("lastrow:")
	print(lastrow)

	sl =  ws.max_row-4

	ws.insert_rows( ws.max_row -3)

	newRowLocation = ws.max_row -4
	print("newRowLocation:")
	print(newRowLocation)

	cell_obj = ws.cell(row=sl, column=1)
	# print(type(cell_obj))
	lastTradeNo = cell_obj.value
	print("Tr.No : " +str(lastTradeNo))

# logList = [tradeType,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,trade_pts]

	# #write to the cell you want, specifying row and column, and value :-)
	ws.cell(column=1,row=newRowLocation, value=(lastTradeNo+1))     #Tr.no

	tradetype = ws.cell(column=2,row=newRowLocation)			 #tradeType
	tradetype.value = logList[0]
	if tradetype.value == 'Short':
		tradetype.fill = PatternFill('solid',fgColor='F4B084')
		tradetype.font = Font(italic = True) 
	else:
		tradetype.fill = PatternFill('solid',fgColor='A9D08E')
		tradetype.font = Font(italic = True)

	ws.cell(column=3,row=newRowLocation, value=logList[1])      	 #date
	ws.cell(column=4,row=newRowLocation, value=logList[2]) 			#day

	entryTime = ws.cell(column=5,row=newRowLocation)			 #entryTime
	entryTime.fill =  PatternFill('solid',fgColor='D6DCE4')
	entryTime.value = logList[3]

	exitTime = ws.cell(column=6,row=newRowLocation)				 #exitTime
	exitTime.fill =  PatternFill('solid',fgColor='D6DCE4') 
	exitTime.value = logList[4]
	
	ws.cell(column=7,row=newRowLocation, value=logList[5])     	 #entry price
	ws.cell(column=8,row=newRowLocation, value=logList[6])       #exit price
	
	tradePoints = ws.cell(column=9,row=newRowLocation)			 #TradePts
	tradePoints.value = logList[7]
	if tradePoints.value <=0 :
		tradePoints.fill = PatternFill('solid',fgColor='FFC7CE')
		tradePoints.font = Font(size = 14,bold = True,color = '9C0006')
	else:
		tradePoints.fill =PatternFill('solid',fgColor='C6EFCE')
		tradePoints.font = Font(size = 14,bold = True,color = '006100')


	totalPnL =  ws.cell(column=9,row=ws.max_row)				#totalPnL
	totalPnL.value ='=sum(I3:I{})'.format(ws.max_row-4)
	
	thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
	for row in ws[newRowLocation:newRowLocation]:  # skip the header
		row.alignment = Alignment(horizontal='center')
		row.border = thin_border

	wb.save(filename=excelpath)
	wb.close()

def shortTradeEntry(shrtOrdrData,excelFile):
	tradeType = 'Short'	
	strPrc = shrtOrdrData[0]["strPrc"]
	dt_object = shrtOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, '%d/%m/%Y %H:%M:%S')
	
	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")

	weekday = entryTimeStamp.strftime("%a")
	trdate = (day+"-"+month)
	
	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = shrtOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)
    
	entry_prc = shrtOrdrData[0]["exePrc"]
	exit_prc = shrtOrdrData[1]["exePrc"]
	

	hedge_cost = shrtOrdrData[0]["hedgeEntryPrc"] - shrtOrdrData[1]["hedgeExitPrc"]
	trade_pts = entry_prc - exit_prc - hedge_cost


	logList = [tradeType,strPrc,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts]

	log_to_excel(excelFile,logList)

def longTradeEntry(longOrdrData,excelFile):
	tradeType = 'Long'
	strPrc = longOrdrData[0]["strPrc"]
	dt_object = longOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, "%d/%m/%Y %H:%M:%S")

	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")

	trdate = (day+"-"+month)

	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")
	weekday = entryTimeStamp.strftime("%a")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = longOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)

	entry_prc = longOrdrData[0]["exePrc"]
	exit_prc = longOrdrData[1]["exePrc"]
	
	hedge_cost = 0
	
	trade_pts = exit_prc - entry_prc
	

	logList = [tradeType,strPrc,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,hedge_cost,trade_pts]

	log_to_excel(excelFile,logList)

def grshortTradeEntry(grshrtOrdrData,excelFile):
	tradeType = 'Short'	
	
	dt_object = grshrtOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, '%d/%m/%Y %H:%M:%S')
	
	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")
	weekday = entryTimeStamp.strftime("%a")
	
	trdate = (day+"-"+month)
	
	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = grshrtOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)
    
	entry_prc = grshrtOrdrData[0]["spotPrc"]
	exit_prc = grshrtOrdrData[1]["spotPrc"]
	
	trade_pts = entry_prc - exit_prc 

	logList = [tradeType,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,trade_pts]

	gr_log_to_excel(excelFile,logList)

def grlongTradeEntry(longOrdrData,excelFile):
	tradeType = 'Long'	
	
	dt_object = longOrdrData[0]["timestamp"]											
	entryTimeStamp = datetime.strptime(dt_object, '%d/%m/%Y %H:%M:%S')
	
	month = entryTimeStamp.strftime("%m")
	day = entryTimeStamp.strftime("%d")

	weekday = entryTimeStamp.strftime("%a")
	trdate = (day+"-"+month)
	
	entry_hour = entryTimeStamp.strftime("%I")
	entry_min = entryTimeStamp.strftime("%M")

	entry_time = (entry_hour+"."+entry_min)

	dt_object2 = longOrdrData[1]["timestamp"]													
	exitTimeStamp = datetime.strptime(dt_object2, "%d/%m/%Y %H:%M:%S")

	exit_hour = exitTimeStamp.strftime("%I")
	exit_min = exitTimeStamp.strftime("%M")

	exit_time = (exit_hour+"."+exit_min)
    
	entry_prc = longOrdrData[0]["spotPrc"]
	exit_prc = longOrdrData[1]["spotPrc"]
	
	trade_pts = exit_prc - entry_prc   

	logList = [tradeType,trdate,weekday,entry_time,exit_time,entry_prc,exit_prc,trade_pts]

	gr_log_to_excel(excelFile,logList)

def settle(sig_data,logbook):
	for order in range(len(sig_data)-1):
		if sig_data[order]["order_type"] == "short":
			shortTradeEntry(sig_data[order:order+2],logbook)	

		if sig_data[order]["order_type"] == "long":
			longTradeEntry(sig_data[order:order+2],logbook)

		if sig_data[order]["order_type"] == "golden_short":
			grshortTradeEntry(sig_data[order:order+2],logbook)	

		if sig_data[order]["order_type"] == "golden_long":
			grlongTradeEntry(sig_data[order:order+2],logbook)

def straddleLog():
	with open(tradelogpath, 'r') as f2:
		data = f2.read()

	data  = data.replace("\n","").split("]")
	d2 = []

	kh = []

	for i in range(0,len(data)-1):
		data[i] = data[i].strip().replace("[","")
		data[i] = " ".join(data[i].split())
		data[i] = ast.literal_eval(data[i])
		d2.append(data[i])
	
	for d in d2:
		kh.append(d)

	settle(kh,logexcelpath)

def goldenratiolog():
	with open(grlogpath, 'r') as f2:
		data = f2.read()

	data  = data.replace("\n","").split("]")
	d2 = []

	kh = []

	for i in range(0,len(data)-1):
		data[i] = data[i].strip().replace("[","")
		data[i] = " ".join(data[i].split())
		data[i] = ast.literal_eval(data[i])
		d2.append(data[i])
	
	for d in d2:
		kh.append(d)

	settle(kh,logexcelpath)


if os.path.isfile(tradelogpath):
	straddleLog()
else:
	print ("No Trades today !")

if os.path.isfile(grlogpath):
	goldenratiolog()
else:
	print ("No Trades today !")




