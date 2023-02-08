# Logs tra

import ast
from calendar import calendar
import calendar
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

from datetime import date, timedelta
from copy import copy
import os

KHexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\KHlog.xlsx"
BYexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\BYlog.xlsx"
# VHexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\VHlog.xlsx"
HVexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\HVlog.xlsx"
# AKexcelpath = "C:\\Users\\amol37007\\OneDrive\\Excel\\Current\\AKlog.xlsx"

acc_names = ['KH', 'BY', 'HV']
excelFiles = [KHexcelpath,BYexcelpath,HVexcelpath]

def log_to_excel(excelpath,logList):
    wb = load_workbook(excelpath)
    ws = wb.worksheets[1]

    lastrow = ws.max_row
    print("lastrow:")
    print(lastrow)

    ws.insert_rows( ws.max_row +1) 
    newRowLocation = ws.max_row +1
    print("newRowLocation:")
    print(newRowLocation)

    sl =  ws.max_row 
    cell_obj = ws.cell(row=sl, column=1)
    print(type(cell_obj))
    lastTradeNo = cell_obj.value
    print("Tr.No : " +str(lastTradeNo))

    # logList = [date,openingbal,dailyPnL,cnTax,cnAmount,withdrawals,additions,closingBal,remarks]

    ws.cell(column=1,row=newRowLocation, value=(lastTradeNo+1))
    ws.cell(column=2,row=newRowLocation,value = logList[0])

    openingBal = ws.cell(column=3,row=newRowLocation)
    openingBal.value = logList[1]
    openingBal.fill =PatternFill('solid',fgColor='C6EFCE')
    openingBal.font = Font(size = 14,bold = True,color = '006100')

    dayPnL = ws.cell(column=4,row=newRowLocation)
    dayPnL.value = logList[2]
    if dayPnL.value <= 0 :
        dayPnL.fill = PatternFill('solid',fgColor='FFC7CE')
        dayPnL.font = Font(size = 14,bold = True,color = '9C0006')
    else:
        dayPnL.fill =PatternFill('solid',fgColor='C6EFCE')
        dayPnL.font = Font(size = 14,bold = True,color = '006100')

    cnTaxAmt = ws.cell(column=5,row=newRowLocation)
    cnTaxAmt.value = logList[3]
    cnTaxAmt.fill =  PatternFill('solid',fgColor='FFC7CE')
    cnTaxAmt.font = Font(size = 14,bold = True,color = '9C0006')

    cnAmt = ws.cell(column=6,row=newRowLocation)
    cnAmt.value = logList[4]
    if cnAmt.value <= 0 :
        cnAmt.fill = PatternFill('solid',fgColor='FFC7CE')
        cnAmt.font = Font(size = 14,bold = True,color = '9C0006')
    else:
        cnAmt.fill =PatternFill('solid',fgColor='C6EFCE')
        cnAmt.font = Font(size = 14,bold = True,color = '006100')

    amtWithdraw = ws.cell(column=7,row=newRowLocation)
    amtWithdraw.value = logList[5]

    amtDeposit = ws.cell(column=8,row=newRowLocation)
    amtDeposit.value = logList[6]

    closingBal = ws.cell(column=9,row=newRowLocation)
    closingBal.value = logList[7]
    closingBal.fill =PatternFill('solid',fgColor='C6EFCE')
    closingBal.font = Font(size = 14,bold = True,color = '006100')

    remark = ws.cell(column=10,row=newRowLocation)
    remark.value = logList[8]

    thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=Side(style='thin'))
    for row in ws[newRowLocation:newRowLocation]:
        row.alignment = Alignment(horizontal='center')
        row.border = thin_border

    wb.save(filename=excelpath)
    wb.close()

def dateCal():    
    today = date.today()
    calendar.day_name[today.weekday()]
    yesterday = today - timedelta(days = 1)
    yes_dayname = calendar.day_name[yesterday.weekday()]

    if yes_dayname == 'Sunday':
        yesterday = (today - timedelta(days = 3))
    else:
        yesterday = (today - timedelta(days = 1))   # change for holidays

    month = str(yesterday.strftime("%b"))
    day = str(yesterday.strftime("%d"))
    year = str(yesterday.strftime("%y"))

    todayDate = (day +' '+ month +' '+ year)

    return todayDate

# logList = [date,openingbal,dailyPnL,cnTax,cnAmount,withdrawals,additions,closingBal,remarks]


def excelDailyEntry(acc_name,withdrawals,additions,remarks,excelFile):
    global prevValue,eodValue
    lstTradeDate = dateCal()

    openingBalDict = {k: v for k, v in prevValue.items() if k.endswith('openingBal')}
    acc_key_openBal = acc_name+'openingBal'
    closingBal = float(openingBalDict[acc_key_openBal])

    pnlDict = {k: v for k, v in eodValue.items() if k.endswith('PnL')}
    pnl_key = acc_name+'PnL'
    pnl = float(pnlDict[pnl_key])   

    closeBalDict = {k: v for k, v in prevValue.items() if k.endswith('closingBal')}
    acc_key_closeBal = acc_name+'closingBal'
    openingBal = float(closeBalDict[acc_key_closeBal])    

    cnAmount = closingBal - openingBal

    cnTax = cnAmount - pnl

    withdrawals = 0 
    additions = 0 
    remarks = 0

    logList = [lstTradeDate,openingBal,pnl,cnTax,cnAmount,withdrawals,additions,closingBal,remarks]
    log_to_excel(excelFile,logList)

def checkQty():
    pass

prevValue = {}
f = open('DailyBalSum.txt','r')
for line in f:
    value = line.split(':')
    prevValue[value[0].strip()] = value[1].strip()

eodValue = {}
file = open('eodSum.txt','r')
for line in file:
    val = line.split(':')
    eodValue[val[0].strip()] = val[1].strip()


withdrawals = 0
additions = 0 
remarks = 0 

for i in range(len(acc_names)):
    excelDailyEntry(acc_names[i],withdrawals,additions,remarks,excelFiles[i])
